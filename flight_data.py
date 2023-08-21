import requests
from auth import KIWIAPI,SEARCH_END_POINT
from datetime import date
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import load_workbook
import copy
import json


HEADER = dict()
HEADER["apikey"] = f"{KIWIAPI}"
LS = []


class FlightData:
    # This class is responsible for structuring the flight data.
    def __init__(self) -> None:
        self.data = []
        self.today = date.today()
        self.six_months = date.today() + relativedelta(months=+6)
        self.temp_dic = dict()

    def update_prices(self, tlv, iata):
        self.fly_from = tlv
        self.fly_to = iata

        d1 = self.today.strftime("%d/%m/%Y")
        d6 = self.six_months.strftime("%d/%m/%Y")

        params = {
            "fly_from": self.fly_from,
            "fly_to": self.fly_to,
            "date_from": d1,
            "date_to ": d6,
            "nights_in_dst_from": 7,
            "nights_in_dst_to": 28,
            "one_for_city": 1,
            "max_stopovers": 0,
            "curr": "USD",
            "flight_type": "round",
            "limit": 1,
        }
        try:
            request = requests.get(url=SEARCH_END_POINT, headers=HEADER, params=params)
            data = request.json()
            print(data)
        except Exception as e:
            print("no flights have been found", e)
        if len(data["data"]) == 0:
            return data
        return data["data"][0]

    def get(self, cites):
        for iata in cites:
            self.fly_from = "TLV"
            self.fly_to = iata

            d1 = self.today.strftime("%d/%m/%Y")
            d6 = self.six_months.strftime("%d/%m/%Y")

            params = {
                "fly_from": self.fly_from,
                "fly_to": self.fly_to,
                "date_from": d1,
                "date_to ": d6,
                "nights_in_dst_from": 7,
                "nights_in_dst_to": 28,
                "one_for_city": 1,
                "max_stopovers": 0,
                "curr": "USD",
                "flight_type": "round",
                "limit": 1,
            }
            request = requests.get(url=SEARCH_END_POINT, headers=HEADER, params=params)
            data = request.json()
            try:
                print(request.status_code)
                if len(data["data"]) == 0:
                    continue
                else:
                    new_data = self.custimize_for_email(data["data"][0])
                    LS.append(copy.copy(new_data))
            except Exception as e:
                print("no flights have been found", e)
        return new_data

    def custimize_for_email(self, data):
        self.temp_dic.clear()

        self.temp_dic["Price"] = data["price"]
        self.temp_dic["Price"] = f'{self.temp_dic["Price"]}$'
        self.temp_dic["Fly_from "] = data["cityCodeFrom"]
        self.temp_dic["Destination"] = data["cityCodeTo"]
        self.temp_dic["Nights"] = data["nightsInDest"]
        self.temp_dic["Date_Flying_out"] = data["local_departure"]
        self.temp_dic["Data_flight_back"] = data["route"][1]["local_departure"]
        self.temp_dic["Link"] = data["deep_link"]

        self.clean_the_date(
            self.temp_dic["Date_Flying_out"],
            self.temp_dic["Data_flight_back"],
        )
        return self.temp_dic

    def clean_the_date(self, local, dest):
        self.temp_dic["Date_Flying_out"] = local.split("T")[0]
        self.temp_dic["Data_flight_back"] = dest.split("T")[0]
        self.temp_dic["hour_Flying_out"] = local.split("T")[1].split(".")[0][0:5]
        self.temp_dic["hour_flight_back"] = dest.split("T")[1].split(".")[0][0:5]

    def make_it_csv(self):
        filename = "deals.csv"
        new_df = pd.DataFrame(LS)
        new_df.to_csv(filename, index=False)

    def web_deals(self):
        # Read the CSV file into a DataFrame
        df = pd.read_csv("deals.csv")

        # Convert the DataFrame to a JSON string
        json_string = df.to_json(orient="records")

        # Convert the JSON string to a Python object
        data = json.loads(json_string)

        # Print the JSON string
        return data

    def convert_to_excel(self):
        # Read the original CSV file into a DataFrame
        df = pd.read_csv("deals.csv")

        # Write the modified DataFrame to a new xlsx file
        df.to_excel("topDeals.xlsx")

        # Modify the width of the columns here
        wb = load_workbook("topDeals.xlsx")
        ws = wb["Sheet1"]
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["J"].width = 20

        for col in ws["H"][1:]:
            temp = col.value
            col.value = "Order"
            col.style = "Hyperlink"
            col.hyperlink = temp
        wb.save("topDeals.xlsx")
